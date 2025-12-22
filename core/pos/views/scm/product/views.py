import json
import pdb
from django.db import transaction
from datetime import datetime
from io import BytesIO
from django.core.exceptions import ValidationError
import xlsxwriter
from django.contrib.auth.mixins import LoginRequiredMixin
from django.db import transaction
from django.http import JsonResponse, HttpResponse, HttpResponseRedirect
from django.shortcuts import render
from django.urls import reverse_lazy
from django.views.generic import CreateView, UpdateView, DeleteView, TemplateView
from django.views.generic.base import View
from openpyxl import load_workbook
from django.db.models import Q
from core.pos.forms import ProductForm, Product, Category
from core.security.mixins import PermissionMixin, ModuleMixin
from django.core.serializers.json import DjangoJSONEncoder


class ProductListView(PermissionMixin, TemplateView):
    template_name = 'scm/product/list.html'
    permission_required = 'view_product'

    def post(self, request, *args, **kwargs):
        data = {}
        action = request.POST.get('action')
        try:
            if action == 'search':
                data = []
                for p in Product.objects.all():
                    data.append(p.toJSON())
            elif action == 'upload_excel':
                archive = request.FILES.get('archive')
                if archive:
                    print("Archivo recibido:", archive.name)
                    with transaction.atomic():
                        workbook = load_workbook(filename=archive, data_only=True)
                        excel = workbook.active
                        for row in range(2, excel.max_row + 1):
                            bar = excel.cell(row=row, column=1).value
                            category_name = excel.cell(row=row, column=2).value
                            name = str(excel.cell(row=row, column=3).value)
                            
                            color = description = excel.cell(row=row, column=4).value
                            publico = description = excel.cell(row=row, column=5).value
                            description = excel.cell(row=row, column=6).value
                            stock = excel.cell(row=row, column=7).value
                            if stock is not None:
                                stock = int(stock)
                            else:
                                stock = 0
                            price = excel.cell(row=row, column=8).value
                            if price is not None:
                                price = float(price)
                            else:
                                price = 0.0
                            pvp = excel.cell(row=row, column=9).value
                            if pvp is not None:
                                pvp = float(pvp)
                            else:
                                pvp = 0.0
                                
                            price_min_sale = excel.cell(row=row, column=10).value
                            if price_min_sale is not None:
                                price_min_sale = float(price_min_sale)
                            else:
                                price_min_sale = 0.0
                                
                            date_into = excel.cell(row=row, column=11).value
                            if isinstance(date_into, str):
                                try:
                                    # Convertir la cadena de fecha al formato adecuado
                                    date_into = datetime.strptime(date_into, '%d/%m/%Y')
                                except ValueError:
                                    # Manejar el caso en que la cadena no sea una fecha válida
                                    print('La cadena no es una fecha válida:', date_into)
                                    date_into = None
                            elif not isinstance(date_into, datetime):
                                # Si no es una cadena ni un objeto datetime, establecerlo como None
                                date_into = None
                                
                            # Validar que el nombre de la categoría no esté vacío
                            if category_name:
                                category, _ = Category.objects.get_or_create(name=category_name)
                                print('ID DE LA CATEGORIA: ', category.id)
                                existing_product = Product.objects.filter(codebar=bar).first()
                                print(existing_product)
                                if existing_product:
                                    existing_product.color = color
                                    existing_product.publico = publico
                                    existing_product.desc = description
                                    existing_product.category = category
                                    existing_product.price = price
                                    existing_product.pvp = pvp
                                    existing_product.price_min_sale = price_min_sale
                                    existing_product.stock += stock
                                    existing_product.save()
                                else:
                                    print('se va a crear')
                                    product = Product.objects.create(
                                        name=name,
                                        color=color,
                                        publico=publico,
                                        desc=description,
                                        codebar=bar,
                                        category=category, 
                                        price=price,
                                        pvp=pvp,
                                        price_min_sale=price_min_sale,
                                        stock=stock,
                                        date_into=date_into
                                    )
                                    print('producto nuevo', product.name)
                            else:
                                # Manejar el caso cuando el nombre de la categoría está ausente
                                print('El nombre de la categoría está ausente en el archivo Excel.')
                                break
                    data['success'] = 'Datos cargados correctamente desde el archivo Excel.'
                else:
                    data['error'] = 'No se ha proporcionado ningún archivo para subir.'
            else:
                data['error'] = 'Acción no válida.'
        except Exception as e:
            print("Error:", e)  # Agregar instrucción de depuración
            data['error'] = str(e)

        return HttpResponse(json.dumps(data, cls=DjangoJSONEncoder), content_type='application/json')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['create_url'] = reverse_lazy('product_create')
        context['title'] = 'Listado de Productos'
        return context

class ProductCreateView(PermissionMixin, CreateView):
    model = Product
    template_name = 'scm/product/create.html'
    form_class = ProductForm
    success_url = reverse_lazy('product_list')
    permission_required = 'add_product'

    def validate_data(self):
        data = {'valid': True}
        try:
            name = self.request.POST['name'].strip()
            category = self.request.POST['category']
            if len(category):
                if Product.objects.filter(name__iexact=name, category_id=category):
                    data['valid'] = False
        except:
            pass
        return JsonResponse(data)

    def post(self, request, *args, **kwargs):
        data = {}
        action = request.POST['action']
        try:
            if action == 'add':
                product = Product()
                product.name = request.POST['name']
                product.color = request.POST['color']
                product.publico = request.POST['publico']
                product.desc = request.POST['desc']
                product.codebar = request.POST['codebar']
                #product.unit = request.POST['unit']
                product.category_id = request.POST['category']
                product.date_into = request.POST['date_into']
                if 'image' in request.FILES:
                    product.image = request.FILES['image']
                product.pvp = float(request.POST['pvp'])
                if product.category.inventoried:
                    product.price = float(request.POST['price'])
                else:
                    product.price = product.pvp
                product.price_min_sale = float(request.POST['price_min_sale'])
                product.save()
            elif action == 'search_category_id':
                data = Category.objects.get(pk=request.POST['id']).toJSON()
            elif action == 'validate_data':
                return self.validate_data()
            else:
                data['error'] = 'No ha seleccionado ninguna opción'
        except Exception as e:
            data['error'] = str(e)
        return HttpResponse(json.dumps(data), content_type='application/json')

    def get_context_data(self, **kwargs):
        context = super().get_context_data()
        context['list_url'] = self.success_url
        context['title'] = 'Nuevo registro de un Producto'
        context['action'] = 'add'
        return context


class ProductUpdateView(PermissionMixin, UpdateView):
    model = Product
    template_name = 'scm/product/create.html'
    form_class = ProductForm
    success_url = reverse_lazy('product_list')
    permission_required = 'change_product'

    def dispatch(self, request, *args, **kwargs):
        self.object = self.get_object()
        return super().dispatch(request, *args, **kwargs)

    def validate_data(self):
        data = {'valid': True}
        try:
            id = self.get_object().id
            name = self.request.POST['name'].strip()
            category = self.request.POST['category']
            if len(category):
                if Product.objects.filter(name__iexact=name, category_id=category).exclude(id=id):
                    data['valid'] = False
        except:
            pass
        return JsonResponse(data)

    def post(self, request, *args, **kwargs):
        data = {}
        action = request.POST['action']
        try:
            if action == 'edit':
                product = self.object
                product.name = request.POST['name']
                product.color = request.POST['color']
                product.publico = request.POST['publico']
                product.desc = request.POST['desc']
                product.codebar = request.POST['codebar']
                product.category_id = request.POST['category']
                product.date_into = request.POST['date_into']
                if 'image-clear' in request.POST:
                    product.remove_image()
                if 'image' in request.FILES:
                    product.image = request.FILES['image']
                product.pvp = float(request.POST['pvp'])
                if product.category.inventoried:
                    product.price = float(request.POST['price'])
                else:
                    product.price = product.pvp
                product.price_min_sale = float(request.POST['price_min_sale'])
                product.save()
            elif action == 'search_category_id':
                data = Category.objects.get(pk=request.POST['id']).toJSON()
            elif action == 'validate_data':
                return self.validate_data()
            else:
                data['error'] = 'No ha seleccionado ninguna opción'
        except Exception as e:
            data['error'] = str(e)
        return HttpResponse(json.dumps(data), content_type='application/json')

    def get_context_data(self, **kwargs):
        context = super().get_context_data()
        context['list_url'] = self.success_url
        context['title'] = 'Edición de un Producto'
        context['action'] = 'edit'
        return context


class ProductDeleteView(PermissionMixin, DeleteView):
    model = Product
    template_name = 'scm/product/delete.html'
    success_url = reverse_lazy('product_list')
    permission_required = 'delete_product'

    def post(self, request, *args, **kwargs):
        data = {}
        try:
            self.get_object().delete()
        except Exception as e:
            data['error'] = str(e)
        return HttpResponse(json.dumps(data), content_type='application/json')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Notificación de eliminación'
        context['list_url'] = self.success_url
        return context


class ProductStockAdjustmentView(ModuleMixin, TemplateView):
    template_name = 'scm/product/stock_adjustment.html'

    def post(self, request, *args, **kwargs):
        data = {}
        action = request.POST['action']
        try:
            if action == 'search_products':
                data = []
                ids = json.loads(request.POST['ids'])
                term = request.POST['term']
                search = Product.objects.filter(category__inventoried=True).exclude(id__in=ids).order_by('name')
                if len(term):
                    search = search.filter(Q(name__icontains=term) | Q(codebar__icontains=term))
                    search = search[0:10]
                for p in search:
                    item = p.toJSON()
                    item['value'] = '{} / {} / {} / {}'.format(p.name, p.category, p.desc, p.codebar)
                    data.append(item)
            elif action == 'create':
                with transaction.atomic():
                    for p in json.loads(request.POST['products']):
                        product = Product.objects.get(pk=p['id'])
                        product.stock = int(p['newstock'])
                        product.save()
            else:
                data['error'] = 'No ha ingresado una opción'
        except Exception as e:
            data['error'] = str(e)
        return HttpResponse(json.dumps(data, cls=DjangoJSONEncoder), content_type='application/json')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Ajuste de Stock de Productos'
        return context


class ProductExportExcelView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        try:
            headers = {
                # 'Código': 15,
                'C.Barra': 15,
                'Categoria': 25,
                'Productos': 25,
                'color':25,
                'publico':25,
                'Descripción': 25,
                'Stock': 15,
                'Precio de Compra': 20,
                'Precio venta': 20,
                'Precio Min Venta': 20,
                'Fecha Ingreso': 10,
            }

            output = BytesIO()
            workbook = xlsxwriter.Workbook(output)
            worksheet = workbook.add_worksheet('productos')
            cell_format = workbook.add_format({'bold': True, 'align': 'center', 'border': 1})
            row_format = workbook.add_format({'align': 'center', 'border': 1})
            index = 0
            for name, width in headers.items():
                worksheet.set_column(first_col=0, last_col=index, width=width)
                worksheet.write(0, index, name, cell_format)
                index += 1
            row = 1
            for product in Product.objects.filter().order_by('id'):
                # worksheet.write(row, 0, product.id, row_format)
                worksheet.write(row, 0, product.codebar, row_format)
                worksheet.write(row, 1, product.category.name, row_format)
                worksheet.write(row, 2, product.name, row_format)
                worksheet.write(row, 3, product.color, row_format)
                worksheet.write(row, 4, product.publico, row_format)
                worksheet.write(row, 5, product.desc, row_format)
                worksheet.write(row, 6, product.stock, row_format)
                worksheet.write(row, 7, format(product.price, '.2f'), row_format)
                worksheet.write(row, 8, format(product.pvp, '.2f'), row_format)
                worksheet.write(row, 9, format(product.price_min_sale, '.2f'), row_format)
                worksheet.write(row, 10, product.date_into.strftime('%d/%m/%Y'), row_format)
                row += 1
            workbook.close()
            output.seek(0)
            response = HttpResponse(output,
                                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename="PRODUCTOS_{}.xlsx"'.format(
                datetime.now().date().strftime('%d_%m_%Y'))
            return response
        except:
            pass
        return HttpResponseRedirect(reverse_lazy('product_list'))


class GeneradorView(ModuleMixin, TemplateView):
    template_name = 'scm/product/qr.html'

    def post(self, request, *args, **kwargs):
        data = {}
        action = request.POST['action']
        try:
            if action == 'search_products':
                data = []
                ids = json.loads(request.POST['ids'])
                term = request.POST['term']
                search = Product.objects.filter(category__inventoried=True).exclude(id__in=ids).order_by('name')
                if len(term):
                    search = search.filter(name__icontains=term)
                    search = search[0:10]
                for p in search:
                    item = p.toJSON()
                    item['value'] = '{} / {}'.format(p.name, p.category.name)
                    data.append(item)
            elif action == 'create':
                with transaction.atomic():
                    for p in json.loads(request.POST['products']):
                        product = Product.objects.get(pk=p['id'])
                        product.stock = int(p['newstock'])
            else:
                data['error'] = 'No ha ingresado una opción'
        except Exception as e:
            data['error'] = str(e)
        return HttpResponse(json.dumps(data), content_type='application/json')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Generador de Productos'
        return context
